from openpyxl import load_workbook
# 加载工作本
wb=load_workbook("t2.xlsx")
# 获取默认的表单
sh1=wb.active
# 读取第一列的前十行
for i in range(1,11):
    print(sh1.cell(i,1).value)
